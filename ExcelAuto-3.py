from pathlib import Path
from openpyxl import load_workbook

# Locate the Excel file in the same folder as this script
file_path = Path(__file__).with_name("hello_world.xlsx")

if not file_path.exists():
    print(f"Error: {file_path} not found. Run ExcelAuto-1.py first.")
else:
    # Load the workbook
    wb = load_workbook(file_path)
    sheet = wb.active

    # Print all rows with their cell values
    print(f"Reading contents of: {file_path.resolve()}")
    for row in sheet.iter_rows(values_only=True):
        print(row)
